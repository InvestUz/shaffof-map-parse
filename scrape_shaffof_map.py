#!/usr/bin/env python3
"""
Harvest document download URLs from https://open-data.mc.uz/map using APZ numbers.

The live map is WebGL; the reliable path is the site's own API after **Tasdiqlash**:
`GET https://open-data.mc.uz/api/apz/details/<id>` (JSON includes `apz_number`, `apz_file`, `project_pdf_file`, …).
When that response is captured, the scraper skips map marker clicks entirely.

Before the UI filter, the scraper may resolve the same **id** via the public list API
`GET https://open-data.mc.uz/api/apz?region_code=<first APZ segment>&apz_number=<full APZ>` (returns `id`, `lat`, `long`, …),
then `GET /api/apz/details/<id>` — no marker click when this succeeds.

Examples:

  # Watch a live browser + slow motion + full-page screenshots of each step:
  python scrape_shaffof_map.py -i dataset.csv -o out.xlsx --trace-actions --slow-mo 200 --debug-screenshots ./shaffof_debug --limit 1

  # Default: bundled Playwright Chromium (no Chrome CDP):
  python scrape_shaffof_map.py --input dataset.csv -o shaffof_urls.xlsx --headless

  # Attach to Chrome with remote debugging (same idea as scrape_invest.py):
  python scrape_shaffof_map.py --input dataset.csv -o out.xlsx --cdp http://127.0.0.1:9222

  # Re-run rows that previously failed (drops failed keys from checkpoint, then resumes):
  python scrape_shaffof_map.py --input dataset.csv -o out.xlsx --resume --retry-errors

  # Re-run rows that were "successful" but have no document URLs (empty checkpoint cells):
  python scrape_shaffof_map.py --input dataset.csv -o out.xlsx --resume --retry-empty

  # Re-run rows that are missing any of the four document URL columns (e.g. Kengash/QSXN still null):
  python scrape_shaffof_map.py --input dataset.csv -o out.xlsx --resume --retry-incomplete-docs

Requires: playwright, pandas, openpyxl (see requirements.txt). Run once: playwright install chromium
"""

from __future__ import annotations

import argparse
import asyncio
import json
import logging
import os
import re
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlencode

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from playwright.async_api import Browser, BrowserContext, Error, Page, Response, async_playwright

from scrape_invest import DEFAULT_CDP, get_browser_context, setup_logging

MAP_URL = "https://open-data.mc.uz/map"
MAX_GOTO_RETRIES = 4
RETRY_DELAY_S = 2.5
CKPT_VERSION = 1

URL_COLS = ("url_art", "url_kengash_xulosasi", "url_ekspertiza", "url_qsxn")
EXTRA_COLS = ("modal_title", "modal_obyekt_nomi", "modal_hududi", "modal_loyihachi", "scrape_error")


@dataclass
class RunDebug:
    """Optional headed / trace mode (set from CLI)."""

    screenshot_dir: Path | None = None
    trace_actions: bool = False


_DBG: RunDebug | None = None


def _set_run_debug(d: RunDebug | None) -> None:
    global _DBG
    _DBG = d


async def _action_trace(page: Page, label: str) -> None:
    """Log a human-readable step and optionally capture a full-page screenshot (for live debugging)."""
    if not _DBG:
        return
    if _DBG.trace_actions:
        logging.info("ACTION: %s", label)
    if _DBG.screenshot_dir:
        _DBG.screenshot_dir.mkdir(parents=True, exist_ok=True)
        safe = re.sub(r"[^\w\u0400-\u04FF\-]+", "_", label).strip("_")[:100] or "step"
        ts = datetime.now().strftime("%H%M%S") + f"_{datetime.now().microsecond // 1000:03d}"
        fn = f"{ts}_{safe}.png"
        try:
            await page.screenshot(path=str(_DBG.screenshot_dir / fn), full_page=True)
        except Error as e:
            logging.warning("Screenshot failed (%s): %s", fn, e)


def _norm_col(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def _norm_apz_match(s: str) -> str:
    """Form / CSV paste often ends with ';' or stray spaces — compare keys after cleanup."""
    return str(s).strip().rstrip(";").strip()


def resolve_apz_column(df: pd.DataFrame) -> str:
    """Find the APZ column (dataset uses 'APZ nomer'; UI text is 'APZ raqami')."""
    mapping = {_norm_col(str(c)): str(c) for c in df.columns}
    for want in ("apz nomer", "apz raqami", "apz number"):
        if want in mapping:
            return mapping[want]
    for norm, orig in mapping.items():
        if "apz" in norm and ("nomer" in norm or "raqami" in norm or "number" in norm):
            return orig
    raise ValueError(
        "Could not find an APZ column. Expected a header like 'APZ nomer' or 'APZ raqami'. "
        f"Got: {list(df.columns)}"
    )


def load_input_csv(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path, sep=";", encoding="utf-8-sig")
    df.columns = [str(c).strip() for c in df.columns]
    return df


def load_checkpoint(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {"version": CKPT_VERSION, "by_apz": {}}
    raw = json.loads(path.read_text(encoding="utf-8"))
    if int(raw.get("version", 0)) != CKPT_VERSION:
        raise ValueError(f"Unsupported checkpoint version in {path}")
    return raw


def save_checkpoint(path: Path, by_apz: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    data = json.dumps({"version": CKPT_VERSION, "by_apz": by_apz}, ensure_ascii=False, indent=0)
    tmp.write_text(data, encoding="utf-8")
    os.replace(tmp, path)


def _ckpt_has_any_doc_url(v: dict[str, Any]) -> bool:
    """True if checkpoint row has at least one harvested http(s) document link."""
    for col in URL_COLS:
        u = v.get(col)
        if isinstance(u, str) and u.strip().startswith("http"):
            return True
    return False


def _ckpt_all_four_doc_urls(v: dict[str, Any]) -> bool:
    """True if every document column has an http(s) link."""
    return _row_has_all_four_doc_urls(v)


def _row_has_all_four_doc_urls(row: dict[str, Any]) -> bool:
    for col in URL_COLS:
        u = row.get(col)
        if not (isinstance(u, str) and u.strip().startswith("http")):
            return False
    return True


def _http_url_or_none(v: Any) -> str | None:
    if isinstance(v, str) and v.strip().startswith("http"):
        return v.strip()
    return None


def _merge_partial_api_rows(a: dict[str, Any] | None, b: dict[str, Any]) -> dict[str, Any]:
    """Prefer b's http URL per column, then a's; merge modal text fields (prefer b when non-empty)."""
    out: dict[str, Any] = {"scrape_error": ""}
    for c in URL_COLS:
        out[c] = _http_url_or_none(b.get(c)) or (_http_url_or_none(a.get(c)) if a else None)
    for c in ("modal_title", "modal_obyekt_nomi", "modal_hududi", "modal_loyihachi"):
        bt = (b.get(c) or "").strip() if isinstance(b.get(c), str) else ""
        at = (a.get(c) or "").strip() if a and isinstance(a.get(c), str) else ""
        out[c] = bt or at
    return out


def _merge_doc_urls_and_modal(
    partial: dict[str, Any] | None,
    urls: dict[str, str | None],
    meta: dict[str, str],
) -> dict[str, Any]:
    """Modal-captured http links win per column; fill gaps from partial API row; modal meta wins text."""
    out: dict[str, Any] = {"scrape_error": ""}
    for c in URL_COLS:
        out[c] = _http_url_or_none(urls.get(c)) or (_http_url_or_none(partial.get(c)) if partial else None)
    for c in ("modal_title", "modal_obyekt_nomi", "modal_hududi", "modal_loyihachi"):
        mt = (meta.get(c) or "").strip()
        pt = (partial.get(c) or "").strip() if partial else ""
        out[c] = mt or pt
    return out


async def _goto_map(page: Page) -> None:
    last: Exception | None = None
    await _action_trace(page, "goto_map_url_start")
    for attempt in range(1, MAX_GOTO_RETRIES + 1):
        try:
            await page.goto(MAP_URL, wait_until="domcontentloaded", timeout=120_000)
            try:
                await page.wait_for_load_state("load", timeout=45_000)
            except Error:
                pass
            await page.wait_for_timeout(2500)
            await page.wait_for_selector(
                "input[placeholder*='APZ'], input[placeholder*='apz' i], .leaflet-container, canvas",
                timeout=45_000,
            )
            await _action_trace(page, "map_shell_ready")
            return
        except Error as e:
            last = e
            logging.warning("goto map attempt %s/%s: %s", attempt, MAX_GOTO_RETRIES, e)
            await asyncio.sleep(RETRY_DELAY_S * attempt)
    raise RuntimeError(f"Failed to load map: {last}")


async def _ensure_filter_visible(page: Page) -> Any:
    """Open the filter drawer if the APZ input is not visible (filter icon / Filtr panel)."""
    apz_input = page.locator('input[placeholder*="APZ"], input[placeholder*="apz" i]').first
    try:
        if await apz_input.is_visible(timeout=1500):
            return apz_input
    except Error:
        pass

    togglers = [
        page.locator("div[class*='animate-filter'] button").first,
        page.locator(".animate-filter button").first,
        page.locator('div.glow button').first,
        page.get_by_role("heading", name=re.compile("filtr", re.I)).locator("xpath=../..").locator("button").first,
    ]
    for loc in togglers:
        try:
            if await loc.count() and await loc.is_visible(timeout=800):
                await loc.click(timeout=5000)
                await page.wait_for_timeout(600)
                if await apz_input.is_visible(timeout=5000):
                    break
        except Error:
            continue

    await apz_input.wait_for(state="visible", timeout=20_000)
    return apz_input


async def _drain_matching_apz_details_from_network(
    page: Page,
    apz: str,
    captured: list[dict[str, Any]],
    *,
    timeout_ms: int = 12_000,
) -> dict[str, Any] | None:
    """Poll buffers filled by page.on('response') for JSON whose apz_number matches this row."""
    want = _norm_apz_match(apz)
    deadline = time.monotonic() + timeout_ms / 1000.0
    while time.monotonic() < deadline:
        for body in captured:
            if _norm_apz_match(str(body.get("apz_number", ""))) == want:
                return body
        await page.wait_for_timeout(200)
    for body in captured:
        if _norm_apz_match(str(body.get("apz_number", ""))) == want:
            return body
    return None


def _apz_details_response_handler(captured: list[dict[str, Any]]):
    async def on_response(response: Response) -> None:
        try:
            if response.status != 200 or "/api/apz/details/" not in response.url:
                return
            ct = (response.headers.get("content-type") or "").lower()
            if "json" not in ct:
                return
            body = await response.json()
            if isinstance(body, dict):
                captured.append(body)
        except Exception:
            return

    return on_response


def _collect_https_urls(obj: Any, sink: list[str]) -> None:
    """Depth-first collect string values that look like absolute http(s) URLs."""
    if isinstance(obj, dict):
        for v in obj.values():
            _collect_https_urls(v, sink)
    elif isinstance(obj, list):
        for x in obj:
            _collect_https_urls(x, sink)
    elif isinstance(obj, str):
        s = obj.strip()
        if s.startswith("http://") or s.startswith("https://"):
            sink.append(s)


def _dedupe_preserve_order(urls: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for u in urls:
        if u not in seen:
            seen.add(u)
            out.append(u)
    return out


def _classify_open_data_detail_urls(j: dict[str, Any]) -> dict[str, str | None]:
    """
    open-data details JSON embeds dx.mc.uz task links: module3 ≈ ART, module2 ≈ Kengash;
    ekspertiza lives on api-ekspertiza.mc.uz (e.g. appeal-final-conclusion-pdf). Do not treat
    project_pdf_file as ekspertiza when it is actually a module2 dx URL.
    """
    raw: list[str] = []
    _collect_https_urls(j, raw)
    ordered = _dedupe_preserve_order(raw)

    art: str | None = None
    keng: str | None = None
    ex: str | None = None
    qsx: str | None = None

    for u in ordered:
        low = u.lower()
        if "api.dx.mc.uz" in low and "/module3/" in low:
            if art is None:
                art = u
            continue
        if "api.dx.mc.uz" in low and "/module2/" in low:
            if keng is None:
                keng = u
            continue
        if "api-ekspertiza.mc.uz" in low or "appeal-final-conclusion" in low:
            if ex is None:
                ex = u
            continue
        if "qsxn" in low or "xabarnoma" in low:
            if qsx is None:
                qsx = u
            continue

    def pick(key: str) -> str | None:
        v = j.get(key)
        return v.strip() if isinstance(v, str) and v.strip().startswith("http") else None

    if art is None:
        art = pick("apz_file")
        if art is None:
            for u in ordered:
                low = u.lower()
                if "art.mc.uz" in low and ("apz" in low or "down-apz" in low or "download" in low):
                    art = u
                    break

    if keng is None:
        keng = pick("kengash_xulosa_file") or pick("kengash_file") or pick("council_file")

    if ex is None:
        ex = pick("ekspertiza_file") or pick("ekspertiza_pdf_file")
        if ex is None:
            pp = pick("project_pdf_file")
            if pp:
                pl = pp.lower()
                if "api.dx.mc.uz" in pl and "/module2/" in pl:
                    if keng is None:
                        keng = pp
                elif "api.dx.mc.uz" in pl and "/module3/" in pl:
                    if art is None:
                        art = pp
                elif "module2" not in pl and "module3" not in pl:
                    ex = pp

    if qsx is None:
        qsx = pick("qsxn_file") or pick("qsxn_xabarnoma_file")

    return {
        "url_art": art,
        "url_kengash_xulosasi": keng,
        "url_ekspertiza": ex,
        "url_qsxn": qsx,
    }


def _row_from_apz_details_api(j: dict[str, Any], expected_apz: str) -> dict[str, Any] | None:
    """
    Map https://open-data.mc.uz/api/apz/details/<id> JSON into our XLSX columns.
    Extra keys may appear over time; unknown file fields stay None.
    """
    if not j:
        return None
    if _norm_apz_match(str(j.get("apz_number", ""))) != _norm_apz_match(str(expected_apz)):
        return None
    name = re.sub(r"\s+", " ", str(j.get("name") or "").strip())
    region = str(j.get("region_name") or "").strip()
    city = str(j.get("city_name") or "").strip()
    hud = ", ".join(x for x in (region, city) if x)
    lo = re.sub(r"\s+", " ", str(j.get("name_design_organization") or "").strip())
    urls = _classify_open_data_detail_urls(j)
    return {
        "url_art": urls["url_art"],
        "url_kengash_xulosasi": urls["url_kengash_xulosasi"],
        "url_ekspertiza": urls["url_ekspertiza"],
        "url_qsxn": urls["url_qsxn"],
        "modal_title": name,
        "modal_obyekt_nomi": name,
        "modal_hududi": hud,
        "modal_loyihachi": lo,
        "scrape_error": "",
    }


REST_APZ_LIST_URL = "https://open-data.mc.uz/api/apz"


async def _fetch_details_json_via_apz_list_api(context: BrowserContext, apz: str) -> dict[str, Any] | None:
    """
    Same records the map uses: GET /api/apz?region_code=…&apz_number=… returns a row with `id`, `lat`, `long`, …
    then GET /api/apz/details/<id> returns full fields (apz_file, project_pdf_file, …).
    """
    want = _norm_apz_match(apz)
    parts = [p.strip() for p in want.split("-") if p.strip()]
    if len(parts) < 2 or not parts[0].isdigit():
        return None
    region_code = parts[0]
    city_code = parts[1] if len(parts) >= 3 and parts[1].isdigit() else None

    variants: list[dict[str, str]] = []
    if city_code:
        variants.append(
            {
                "page": "1",
                "per_page": "500",
                "region_code": region_code,
                "city_code": city_code,
                "apz_number": want,
            }
        )
    variants.append(
        {"page": "1", "per_page": "500", "region_code": region_code, "apz_number": want},
    )

    chosen: dict[str, Any] | None = None
    for params in variants:
        url = f"{REST_APZ_LIST_URL}?{urlencode(params)}"
        try:
            resp = await context.request.get(url, timeout=45_000)
            if resp.status != 200:
                logging.debug("apz list API %s -> HTTP %s", url, resp.status)
                continue
            data = await resp.json()
        except Exception as e:
            logging.debug("apz list API request failed %s: %s", url, e)
            continue
        if not isinstance(data, list):
            continue
        for row in data:
            if not isinstance(row, dict):
                continue
            if _norm_apz_match(str(row.get("apz_number", ""))) != want:
                continue
            chosen = row
            break
        if chosen is not None:
            break

    if not chosen:
        return None
    oid = chosen.get("id")
    if oid is None:
        return None
    detail_url = f"https://open-data.mc.uz/api/apz/details/{int(oid)}"
    try:
        r2 = await context.request.get(detail_url, timeout=45_000)
        if r2.status != 200:
            logging.debug("apz details API %s -> HTTP %s", detail_url, r2.status)
            return None
        detail = await r2.json()
    except Exception as e:
        logging.debug("apz details API failed %s: %s", detail_url, e)
        return None
    if isinstance(detail, dict):
        logging.info("Resolved APZ via REST /api/apz list (id=%s) then /api/apz/details", oid)
        return detail
    return None


async def _apply_apz_filter(page: Page, apz: str) -> dict[str, Any] | None:
    apz_input = await _ensure_filter_visible(page)
    await _action_trace(page, "filter_drawer_apz_field_visible")
    await apz_input.click(timeout=5000)
    await apz_input.fill("")
    await apz_input.fill(apz.strip())
    await _action_trace(page, f"filter_filled_apz_{apz.strip()[:40]}")

    confirm = page.get_by_role("button", name=re.compile(r"^\s*Tasdiqlash\s*$", re.I))
    captured: list[dict[str, Any]] = []

    on_response = _apz_details_response_handler(captured)

    matched: dict[str, Any] | None = None
    page.on("response", on_response)
    try:
        await confirm.click(timeout=15_000)
        await _action_trace(page, "filter_clicked_Tasdiqlash")
        deadline = time.monotonic() + 14.0
        want = _norm_apz_match(apz)
        while time.monotonic() < deadline:
            for body in captured:
                if _norm_apz_match(str(body.get("apz_number", ""))) == want:
                    matched = body
                    break
            if matched is not None:
                break
            await page.wait_for_timeout(200)
        if matched is None:
            for body in captured:
                if _norm_apz_match(str(body.get("apz_number", ""))) == want:
                    matched = body
                    break
    finally:
        try:
            page.remove_listener("response", on_response)
        except Exception:
            pass

    if matched is not None:
        logging.info("Captured open-data API /api/apz/details for APZ %s (skip map marker)", want)
        return matched

    # No details payload yet — wait for map, then fall back to marker / canvas strategies.
    await page.wait_for_timeout(4000)
    for _ in range(3):
        try:
            await page.keyboard.press("Escape")
            await page.wait_for_timeout(400)
        except Error:
            pass
    await _action_trace(page, "filter_panel_dismissed_escape")
    return None


_MARKER_WAIT_JS = """() => {
  /* Yandex Maps JSAPI: placemarks are custom <ymaps> nodes (e.g. class *-svg-icon), not Leaflet. */
  if (document.querySelector('ymaps[class*="-svg-icon"], ymaps[class*="svg-icon"]')) return true;
  const pane = document.querySelector('.leaflet-marker-pane');
  if (pane && pane.children.length) return true;
  if (document.querySelector('.maplibregl-marker, .mapboxgl-marker')) return true;
  if (document.querySelector('.leaflet-marker-icon, .leaflet-interactive')) return true;
  if (document.querySelector('.leaflet-overlay-pane svg path')) return true;
  return false;
}"""

_MARKER_CLICK_JS = """() => {
  function visible(el) {
    if (!el || el.nodeType !== 1) return false;
    const s = window.getComputedStyle(el);
    if (s.display === 'none' || s.visibility === 'hidden' || Number(s.opacity) === 0) return false;
    const r = el.getBoundingClientRect();
    return r.width >= 1 && r.height >= 1;
  }
  function firePointerClick(el) {
    const r = el.getBoundingClientRect();
    const x = r.left + r.width / 2;
    const y = r.top + r.height / 2;
    const opts = { bubbles: true, cancelable: true, view: window, clientX: x, clientY: y };
    try {
      el.dispatchEvent(new PointerEvent('pointerdown', { ...opts, pointerId: 1, pointerType: 'mouse' }));
    } catch (e) {
      /* older engines */
    }
    el.dispatchEvent(new MouseEvent('mousedown', opts));
    el.dispatchEvent(new MouseEvent('mouseup', opts));
    el.dispatchEvent(new MouseEvent('click', opts));
    if (typeof el.click === 'function') try { el.click(); } catch (e2) {}
    return true;
  }
  const tryList = (nodes) => {
    for (const el of nodes) {
      if (visible(el) && firePointerClick(el)) return el.tagName + '.' + (el.className || '').toString().slice(0, 40);
    }
    return null;
  };
  const selectors = [
    'ymaps[class*="-svg-icon"]',
    'ymaps[class*="svg-icon"]',
    '.leaflet-marker-pane img',
    '.leaflet-marker-pane > div',
    '.leaflet-marker-pane svg',
    '.leaflet-marker-pane .leaflet-marker-icon',
    '.leaflet-shadow-pane img',
    '.leaflet-overlay-pane svg path.leaflet-interactive',
    '.leaflet-overlay-pane path[stroke-width]',
    '.maplibregl-marker',
    '.mapboxgl-marker',
    '[class*="maplibregl-marker"]',
    '[class*="mapboxgl-marker"]',
  ];
  for (const sel of selectors) {
    const hit = tryList(document.querySelectorAll(sel));
    if (hit) return 'js:' + sel + ':' + hit;
  }
  const pane = document.querySelector('.leaflet-marker-pane');
  if (pane) {
    const hit = tryList(pane.querySelectorAll('img, div, svg, button, span'));
    if (hit) return 'js:pane-any:' + hit;
  }
  return null;
}"""


async def _wait_marker_layer(page: Page, *, timeout_ms: int = 28_000) -> None:
    """Wait until the map paints at least one marker-like layer (best-effort)."""
    try:
        await page.wait_for_function(_MARKER_WAIT_JS, timeout=timeout_ms)
    except Error:
        logging.debug("marker wait_for_function timed out (continuing with fallbacks)")


async def _click_map_area_center(page: Page) -> bool:
    """Last resort: click the middle of the main map container."""
    for sel in (".leaflet-container", ".maplibregl-map", "[class*='maplibregl']", "canvas.maplibregl-canvas"):
        loc = page.locator(sel).first
        try:
            if await loc.count() == 0:
                continue
            box = await loc.bounding_box()
            if not box or box["width"] < 20 or box["height"] < 20:
                continue
            await page.mouse.click(box["x"] + box["width"] / 2, box["y"] + box["height"] / 2)
            return True
        except Error:
            continue
    return False


_IS_CANVAS_ONLY_MAP_JS = """() => {
  const canvas = document.querySelector("canvas");
  if (!canvas) return false;
  const r = canvas.getBoundingClientRect();
  if (r.width < 80 || r.height < 80) return false;
  /* Yandex layer: pin is a real <ymaps> target — do not treat as WebGL-only (avoids wheel zoom + grid). */
  if (document.querySelector('ymaps[class*="-svg-icon"], ymaps[class*="svg-icon"]')) return false;
  const pane = document.querySelector(".leaflet-marker-pane");
  if (pane && pane.children && pane.children.length) return false;
  return !document.querySelector(".leaflet-container");
}"""


async def _dismiss_modal_if_open(page: Page) -> None:
    dlg = page.locator('[role="dialog"]').first
    try:
        if await dlg.is_visible(timeout=250):
            await page.keyboard.press("Escape")
            await page.wait_for_timeout(300)
    except Error:
        pass


async def _object_ui_opened(page: Page) -> bool:
    """Modal may use role=dialog or a sheet without that role; object card uses a known image alt."""
    try:
        if await page.locator('[role="dialog"]').first.is_visible(timeout=120):
            return True
    except Error:
        pass
    try:
        if await page.locator('img[alt="object image"], img[alt*="object" i]').first.is_visible(timeout=120):
            return True
    except Error:
        pass
    try:
        open_sheet = page.locator('[data-state="open"]').filter(has=page.locator("h5"))
        if await open_sheet.count() and await open_sheet.first.is_visible(timeout=120):
            return True
    except Error:
        pass
    try:
        if await page.locator("h5").first.is_visible(timeout=120):
            t = await page.locator("h5").first.inner_text()
            if t and len(t.strip()) > 15:
                return True
    except Error:
        pass
    return False


async def _try_click_ymaps_placemark(page: Page) -> bool:
    """Yandex Maps JSAPI renders filtered pins as <ymaps class='...-svg-icon' style='position:absolute;…'>."""
    loc = page.locator("ymaps[class*='-svg-icon'], ymaps[class*='svg-icon']").first
    try:
        await loc.wait_for(state="visible", timeout=12_000)
        await loc.scroll_into_view_if_needed(timeout=5000)
        await loc.click(timeout=12_000, force=True)
        await page.wait_for_timeout(500)
        if await _object_ui_opened(page):
            logging.info("Opened object UI via Yandex <ymaps> svg-icon placemark click")
            return True
    except Error as e:
        logging.debug("ymaps placemark click: %s", e)
    return False


async def _try_open_object_via_apz_text(page: Page, apz: str) -> bool:
    """After filtering, some layouts expose the hit as a list/card row (not a map canvas hit-target)."""
    apz = apz.strip()
    if not apz:
        return False
    esc = re.escape(apz)
    pat = re.compile(esc)
    for role in ("button", "link", "cell"):
        loc = page.get_by_role(role, name=pat)
        try:
            if await loc.count() == 0:
                continue
            await loc.first.click(timeout=5000, force=True)
            await page.wait_for_timeout(400)
            if await _object_ui_opened(page):
                logging.info("Opened object UI via %s matching APZ text", role)
                return True
        except Error:
            continue
    try:
        loc = page.locator("div.cursor-pointer, [class*='cursor-pointer']").filter(has_text=pat).first
        if await loc.count():
            await loc.click(timeout=5000, force=True)
            await page.wait_for_timeout(400)
            if await _object_ui_opened(page):
                logging.info("Opened object UI via cursor-pointer row with APZ")
                return True
    except Error:
        pass
    return False


async def _largest_canvas_locator(page: Page) -> Any | None:
    best: tuple[float, Any] | None = None
    n = await page.locator("canvas").count()
    for i in range(n):
        loc = page.locator("canvas").nth(i)
        try:
            box = await loc.bounding_box()
            if not box:
                continue
            area = box["width"] * box["height"]
            if area < 2000:
                continue
            if best is None or area > best[0]:
                best = (area, loc)
        except Error:
            continue
    return best[1] if best else page.locator("canvas").first


async def _nudge_zoom_on_canvas(page: Page) -> None:
    """Zoom in slightly so a lone filtered pin is easier to hit-test on WebGL maps."""
    try:
        loc = await _largest_canvas_locator(page)
        box = await loc.bounding_box()
        if not box:
            return
        cx, cy = box["x"] + box["width"] / 2, box["y"] + box["height"] / 2
        await page.mouse.move(cx, cy)
        for _ in range(4):
            await page.mouse.wheel(0, -400)
            await page.wait_for_timeout(350)
    except Error:
        pass


_MAPBOX_TRY_CLICK_JS = r"""() => {
  function findMap(el) {
    let cur = el;
    for (let i = 0; i < 10 && cur; i++, cur = cur.parentElement) {
      const m = cur._map || cur.__map;
      if (m && typeof m.queryRenderedFeatures === "function" && typeof m.getCanvas === "function") return m;
    }
    return null;
  }
  const starters = Array.from(document.querySelectorAll(".maplibregl-map, .mapboxgl-map, canvas"));
  for (const el of starters) {
    const m = findMap(el);
    if (!m) continue;
    const canvas = m.getCanvas();
    if (!canvas) continue;
    const w = canvas.width || canvas.clientWidth;
    const h = canvas.height || canvas.clientHeight;
    if (w < 32 || h < 32) continue;
    for (let gy = 0.35; gy <= 0.65; gy += 0.05) {
      for (let gx = 0.35; gx <= 0.65; gx += 0.05) {
        const x = gx * w;
        const y = gy * h;
        let feats = [];
        try {
          feats = m.queryRenderedFeatures([x, y]) || [];
        } catch (e) {
          continue;
        }
        if (!feats.length) continue;
        try {
          const lngLat = m.unproject([x, y]);
          m.fire("click", { point: { x, y }, lngLat });
        } catch (e2) {
          /* ignore */
        }
        return "ok";
      }
    }
  }
  return null;
}"""


async def _try_maplibre_mapbox_programmatic_click(page: Page) -> bool:
    """If MapLibre/Mapbox exposes _map on a container, query rendered features and fire a map click."""
    try:
        hit = await page.evaluate(_MAPBOX_TRY_CLICK_JS)
        if not hit:
            return False
        logging.info("Map GL programmatic click: %s", hit)
        await page.wait_for_timeout(900)
        return await _object_ui_opened(page)
    except Error as e:
        logging.debug("map GL programmatic click: %s", e)
        return False


async def _click_canvas_grid_for_modal(page: Page, *, steps: int = 6, pause_ms: int = 80) -> bool:
    """
    open-data.mc.uz uses a WebGL <canvas> map (no Leaflet DOM markers). After APZ filter,
    the pin is hit-tested on the canvas — we search a grid of click points until the object dialog opens.
    """
    loc = await _largest_canvas_locator(page)
    await loc.wait_for(state="attached", timeout=25_000)
    box = await loc.bounding_box()
    if not box or box["width"] < 40 or box["height"] < 40:
        return False
    x0, y0, w, h = box["x"], box["y"], box["width"], box["height"]
    mx, my = w * 0.06, h * 0.06
    iw, ih = w - 2 * mx, h - 2 * my
    n = max(5, min(int(steps), 11))
    for yi in range(n):
        for xi in range(n):
            await _dismiss_modal_if_open(page)
            rel_x = mx + (xi + 0.5) * (iw / n)
            rel_y = my + (yi + 0.5) * (ih / n)
            await page.mouse.click(x0 + rel_x, y0 + rel_y)
            await page.wait_for_timeout(pause_ms)
            if await _object_ui_opened(page):
                return True
    return False


async def _click_first_marker(page: Page, apz: str) -> None:
    """
    Leaflet DOM markers, list row with APZ text, or (for open-data.mc.uz) a WebGL canvas map — use a
    canvas click grid when there is no marker pane.
    """
    await _action_trace(page, "marker_step_start")
    if await _try_open_object_via_apz_text(page, apz):
        await _action_trace(page, "marker_opened_via_apz_text_row")
        return

    if await _try_click_ymaps_placemark(page):
        await _action_trace(page, "marker_opened_via_ymaps_svg_icon")
        return

    canvas_only = False
    try:
        canvas_only = bool(await page.evaluate(_IS_CANVAS_ONLY_MAP_JS))
    except Error:
        pass

    if canvas_only:
        logging.info("Canvas/WebGL map (no DOM markers); searching canvas for filtered pin…")
        if await _click_canvas_grid_for_modal(page):
            return
        logging.info("Zooming in on map and re-trying canvas hit search…")
        await _nudge_zoom_on_canvas(page)
        if await _click_canvas_grid_for_modal(page):
            return
        if await _try_maplibre_mapbox_programmatic_click(page):
            return
        logging.debug("Primary canvas grid missed; skipping Leaflet marker-pane wait")
    else:
        await _wait_marker_layer(page, timeout_ms=28_000)

    try:
        tag = await page.evaluate(_MARKER_CLICK_JS)
        if tag:
            logging.info("Marker click via evaluate: %s", tag)
            await page.wait_for_timeout(600)
            return
    except Error as e:
        logging.debug("marker evaluate click: %s", e)

    selectors = [
        "ymaps[class*='-svg-icon']",
        "ymaps[class*='svg-icon']",
        ".leaflet-marker-pane img",
        ".leaflet-marker-pane > div",
        ".leaflet-marker-pane svg",
        ".leaflet-marker-icon",
        ".leaflet-interactive",
        ".leaflet-overlay-pane svg path.leaflet-interactive",
        ".mapboxgl-marker",
        "div.maplibregl-marker",
        "[class*='mapboxgl-marker']",
        "img.leaflet-marker-icon",
    ]
    last_err: Exception | None = None
    for sel in selectors:
        loc = page.locator(sel).first
        try:
            await loc.wait_for(state="attached", timeout=3500)
            await loc.scroll_into_view_if_needed(timeout=5000)
            await loc.click(timeout=12_000, force=True)
            logging.info("Marker click via locator: %s", sel)
            return
        except Error as e:
            last_err = e
            continue

    if await _click_map_area_center(page):
        logging.info("Marker fallback: clicked map container center")
        await page.wait_for_timeout(600)
        try:
            await page.locator('[role="dialog"]').first.wait_for(state="visible", timeout=4000)
            return
        except Error:
            pass

    if await _click_canvas_grid_for_modal(page):
        logging.info("Opened object modal via canvas grid (fallback)")
        return

    if await _try_maplibre_mapbox_programmatic_click(page):
        return

    await _action_trace(page, "marker_all_strategies_failed")
    raise RuntimeError(f"No clickable map marker found (DOM + canvas grid failed): {last_err}")


async def _wait_object_modal(page: Page) -> Any:
    """Object sheet may be role=dialog or a Radix/shadcn panel with data-state=open."""
    for sel in ('[role="dialog"]', '[data-state="open"]'):
        loc = page.locator(sel).first
        try:
            await loc.wait_for(state="visible", timeout=10_000)
            if await loc.locator("h5").count():
                return loc
        except Error:
            continue
    await page.locator('img[alt="object image"], img[alt*="object" i]').first.wait_for(
        state="visible", timeout=15_000
    )
    return page.locator('[data-state="open"], [role="dialog"]').first


async def _wait_modal_content_hydrated(page: Page, modal: Any, *, timeout_ms: int = 25_000) -> None:
    """Wait out React placeholders like 'ART (undefined)' after the sheet shell is visible."""
    deadline = time.monotonic() + timeout_ms / 1000.0
    while time.monotonic() < deadline:
        try:
            if not await modal.is_visible():
                return
            t = (await modal.inner_text()).lower()
            if "undefined" not in t:
                return
        except Error:
            return
        await page.wait_for_timeout(280)


def _is_garbage_modal_value(t: str) -> bool:
    s = re.sub(r"\s+", " ", (t or "").strip())
    if not s or s in (",", ":", "-", "—", "____", "_____"):
        return True
    return bool(re.fullmatch(r"[\s,:\-—_]+", s))


def _text_is_mostly_label(label_regex: str, text: str) -> bool:
    """True when the cell text is essentially the field label (no separate value yet)."""
    t = re.sub(r"\s+", " ", text.strip())
    if len(t) > 120:
        return False
    return bool(re.fullmatch(rf"\s*(?:{label_regex})\s*:?\s*", t, re.I))


async def _read_modal_kv(modal: Any) -> dict[str, str]:
    out: dict[str, str] = {}
    try:
        out["modal_title"] = re.sub(r"\s+", " ", (await modal.locator("h5").first.inner_text()).strip())
    except Error:
        out["modal_title"] = ""

    async def pair_field(key: str, label_regexes: tuple[str, ...]) -> None:
        if out.get(key):
            return
        for label_pat in label_regexes:
            try:
                lab = modal.locator("p, span, label, div").filter(has_text=re.compile(label_pat, re.I))
                if await lab.count() == 0:
                    continue
                row = lab.first.locator("xpath=ancestor::div[contains(@class,'flex')][1]")
                if await row.count() == 0:
                    row = lab.first.locator("xpath=ancestor::*[contains(@class,'flex')][1]")
                kids = row.locator(":scope > *")
                nk = await kids.count()
                if nk >= 2:
                    for idx in range(1, nk):
                        t = re.sub(r"\s+", " ", (await kids.nth(idx).inner_text()).strip())
                        if _is_garbage_modal_value(t):
                            continue
                        if any(_text_is_mostly_label(lp, t) for lp in label_regexes):
                            continue
                        out[key] = t
                        return
                vals = row.locator("p.text-sm, p[class*='text-sm'], div[class*='text-sm'], span, div.font-medium")
                for j in range(await vals.count()):
                    t = re.sub(r"\s+", " ", (await vals.nth(j).inner_text()).strip())
                    if _is_garbage_modal_value(t):
                        continue
                    if any(_text_is_mostly_label(lp, t) for lp in label_regexes):
                        continue
                    if not re.search(label_pat, t, re.I):
                        out[key] = t
                        return
                    if len(t) > len(label_pat) + 4:
                        out[key] = t
                        return
            except Error:
                continue

    await pair_field(
        "modal_obyekt_nomi",
        (r"Obyekt\s+Nomi", r"^\s*Nomi\s*:?", r"\bNomi\s*:"),
    )
    await pair_field("modal_hududi", (r"^\s*Hududi\s*:?", r"\bHududi\s*:"))
    await pair_field("modal_loyihachi", (r"^\s*Loyihachi\s*:?", r"\bLoyihachi\s*:"))
    return out


async def _capture_url_from_doc_button(page: Page, button: Any) -> str | None:
    """Document buttons may open a new tab, navigate the same page, or trigger a download."""
    try:
        async with page.context.expect_page(timeout=18_000) as pg_info:
            await button.click(timeout=10_000)
        new_page = await pg_info.value
        try:
            await new_page.wait_for_load_state("domcontentloaded", timeout=25_000)
        except Error:
            pass
        url = new_page.url
        await new_page.close()
        if url and url != "about:blank":
            return url
    except Error:
        pass

    try:
        async with page.expect_popup(timeout=18_000) as pop_info:
            await button.click(timeout=10_000)
        pop = await pop_info.value
        try:
            await pop.wait_for_load_state("domcontentloaded", timeout=25_000)
        except Error:
            pass
        url = pop.url
        await pop.close()
        if url and url != "about:blank":
            return url
    except Error:
        pass

    prev = page.url
    try:
        async with page.expect_navigation(timeout=18_000, wait_until="domcontentloaded"):
            await button.click(timeout=10_000)
        if page.url != prev and "open-data.mc.uz/map" not in page.url:
            u = page.url
            await page.go_back(wait_until="domcontentloaded", timeout=30_000)
            await _wait_object_modal(page)
            return u
    except Error:
        pass

    try:
        async with page.expect_download(timeout=18_000) as dl_info:
            await button.click(timeout=10_000)
        dl = await dl_info.value
        u = dl.url
        if u:
            return u
    except Error:
        pass

    return None


def _ekspertiza_pdf_response_handler(sink: list[str]):
    """Collect PDF response URLs from api-ekspertiza.mc.uz (internal id, not reestr number)."""

    async def on_response(response: Response) -> None:
        try:
            if response.status != 200:
                return
            u = response.url
            if "api-ekspertiza.mc.uz" not in u.lower():
                return
            ct = (response.headers.get("content-type") or "").lower()
            if "pdf" not in ct and "octet-stream" not in ct:
                return
            if u not in sink:
                sink.append(u)
        except Exception:
            return

    return on_response


async def _extract_doc_urls(page: Page, modal: Any) -> dict[str, str | None]:
    """
    Map document line labels to URL columns.
    Order follows typical modal listing: ART, Kengash, Ekspertiza, QSXN.
    Uses the same sheet root as _read_modal_kv (dialog or data-state=open), not only role=dialog.
    """
    patterns: list[tuple[str, re.Pattern[str]]] = [
        ("url_art", re.compile(r"\bART\b", re.I)),
        ("url_kengash_xulosasi", re.compile(r"Kengash", re.I)),
        (
            "url_ekspertiza",
            re.compile(r"Ekspertiza(?:\s+Reestr)?|Expertiza", re.I),
        ),
        ("url_qsxn", re.compile(r"QSXN", re.I)),
    ]
    out: dict[str, str | None] = {key: None for key, _ in patterns}

    for key, pat in patterns:
        try:
            await modal.wait_for(state="visible", timeout=8000)
        except Error:
            break
        label_cell = modal.locator("p, span, div").filter(has_text=pat)
        row = modal.locator("div.flex.justify-between").filter(has=label_cell)
        if await row.count() == 0:
            row = modal.locator("div.flex").filter(has=label_cell)
        r0 = row.first
        try:
            link = r0.locator("a[href^='http']").first
            if await link.count():
                href = (await link.get_attribute("href") or "").strip()
                if href.startswith("http"):
                    out[key] = href
                    continue
            btn = r0.locator("button").first
            if await btn.count() == 0:
                continue
            url = await _capture_url_from_doc_button(page, btn)
            out[key] = url
        except Error as e:
            logging.debug("doc row %s: %s", key, e)
            out[key] = None

    await _fill_doc_urls_from_modal_link_scan(modal, out)
    return out


async def _fill_doc_urls_from_modal_link_scan(modal: Any, out: dict[str, str | None]) -> None:
    """Row-based matching often misses Ekspertiza/QSXN; scan the sheet for known hosts."""
    try:
        if not out.get("url_ekspertiza"):
            loc = modal.locator(
                'a[href*="api-ekspertiza.mc.uz"], a[href*="ekspertiza.mc.uz"], a[href*="new-ekspertiza"]'
            )
            n = await loc.count()
            for i in range(min(n, 12)):
                href = (await loc.nth(i).get_attribute("href") or "").strip()
                if not href.startswith("http"):
                    continue
                low = href.lower()
                if any(
                    x in low
                    for x in ("appeal-final-conclusion", "final-conclusion-pdf", "conclusion-pdf", "/pdf/")
                ):
                    out["url_ekspertiza"] = href
                    break
            if not out.get("url_ekspertiza") and n:
                href = (await loc.first.get_attribute("href") or "").strip()
                if href.startswith("http"):
                    out["url_ekspertiza"] = href

        if not out.get("url_qsxn"):
            locq = modal.locator('a[href*="qsxn"], a[href*="xabarnoma"], a[href*="dx.mc.uz"][href*="module1"]')
            nq = await locq.count()
            for i in range(min(nq, 8)):
                href = (await locq.nth(i).get_attribute("href") or "").strip()
                if href.startswith("http"):
                    out["url_qsxn"] = href
                    break
    except Error as e:
        logging.debug("modal link scan: %s", e)


def _empty_scrape_result(message: str) -> dict[str, Any]:
    return {
        **{c: None for c in URL_COLS},
        "modal_title": "",
        "modal_obyekt_nomi": "",
        "modal_hududi": "",
        "modal_loyihachi": "",
        "scrape_error": message,
    }


async def scrape_one_apz(page: Page, apz: str) -> dict[str, Any]:
    errors: list[str] = []
    slug = re.sub(r"[^\w\-]+", "_", apz.strip())[:50] or "apz"
    await _action_trace(page, f"row_start_{slug}")
    await _goto_map(page)
    await _action_trace(page, "after_goto_map")

    partial_row: dict[str, Any] | None = None
    try:
        rest_json = await _fetch_details_json_via_apz_list_api(page.context, apz)
    except Exception as e:
        logging.debug("REST /api/apz + details prefetch: %s", e)
        rest_json = None
    rest_row = _row_from_apz_details_api(rest_json, apz) if rest_json else None
    if rest_row is not None:
        if _row_has_all_four_doc_urls(rest_row):
            await _action_trace(page, "row_complete_via_rest_list_then_details")
            return rest_row
        partial_row = rest_row
        logging.info(
            "open-data /api/apz/details returned fewer than four file URLs for %s "
            "(Ekspertiza/QSXN are often only reestr numbers in JSON); continuing through map/modal.",
            apz.strip()[:72],
        )

    details: dict[str, Any] | None = None
    try:
        details = await _apply_apz_filter(page, apz)
    except Error as e:
        errors.append(f"filter: {e}")
        await _action_trace(page, "filter_failed")
        if partial_row:
            return {**{c: partial_row.get(c) for c in URL_COLS}, **{k: partial_row.get(k, "") for k in EXTRA_COLS if k != "scrape_error"}, "scrape_error": " | ".join(errors)}
        return _empty_scrape_result(" | ".join(errors))

    api_row = _row_from_apz_details_api(details, apz) if details else None
    if api_row is not None:
        if _row_has_all_four_doc_urls(api_row):
            await _action_trace(page, "row_complete_via_api_details_json")
            return api_row
        partial_row = _merge_partial_api_rows(partial_row, api_row)

    await _action_trace(page, "after_apply_filter_before_marker")
    captured_after: list[dict[str, Any]] = []
    on_after = _apz_details_response_handler(captured_after)
    page.on("response", on_after)
    modal = None
    try:
        try:
            await _click_first_marker(page, apz)
        except Exception as e:
            errors.append(f"marker: {e}")
            await _action_trace(page, "marker_step_failed")
            if partial_row:
                return {**{c: partial_row.get(c) for c in URL_COLS}, **{k: partial_row.get(k, "") for k in EXTRA_COLS if k != "scrape_error"}, "scrape_error": " | ".join(errors)}
            return _empty_scrape_result(" | ".join(errors))

        await _action_trace(page, "after_marker_before_modal_wait")
        try:
            modal = await _wait_object_modal(page)
        except Error as e:
            errors.append(f"modal: {e}")
            await _action_trace(page, "modal_wait_failed")

        matched_after = await _drain_matching_apz_details_from_network(
            page, apz, captured_after, timeout_ms=14_000
        )
    finally:
        try:
            page.remove_listener("response", on_after)
        except Exception:
            pass

    if matched_after:
        api_row2 = _row_from_apz_details_api(matched_after, apz)
        if api_row2 is not None:
            if _row_has_all_four_doc_urls(api_row2):
                try:
                    await page.keyboard.press("Escape")
                    await page.wait_for_timeout(400)
                except Error:
                    pass
                await _action_trace(page, "row_complete_via_api_after_marker")
                return api_row2
            partial_row = _merge_partial_api_rows(partial_row, api_row2)

    if modal is None:
        if partial_row:
            return {**{c: partial_row.get(c) for c in URL_COLS}, **{k: partial_row.get(k, "") for k in EXTRA_COLS if k != "scrape_error"}, "scrape_error": " | ".join(errors) if errors else ""}
        return _empty_scrape_result(" | ".join(errors))

    eks_pdf_urls: list[str] = []
    on_eks_pdf = _ekspertiza_pdf_response_handler(eks_pdf_urls)
    page.on("response", on_eks_pdf)
    try:
        await _wait_modal_content_hydrated(page, modal)
        meta = await _read_modal_kv(modal)
        urls = await _extract_doc_urls(page, modal)
    finally:
        try:
            page.remove_listener("response", on_eks_pdf)
        except Exception:
            pass
    if not urls.get("url_ekspertiza") and eks_pdf_urls:
        urls["url_ekspertiza"] = eks_pdf_urls[-1]
        logging.info("Ekspertiza URL taken from network PDF response: %s", urls["url_ekspertiza"][:80])

    try:
        await page.keyboard.press("Escape")
        await page.wait_for_timeout(400)
    except Error:
        pass

    row_out = _merge_doc_urls_and_modal(partial_row, urls, meta)
    await _action_trace(page, "row_complete_urls_extracted")
    return row_out


def merge_checkpoint_into_df(df: pd.DataFrame, apz_col: str, ckpt: dict[str, Any]) -> pd.DataFrame:
    by_apz = ckpt.get("by_apz") or {}
    extra = list(URL_COLS) + list(EXTRA_COLS)
    for c in extra:
        if c not in df.columns:
            df[c] = None
    for i, r in df.iterrows():
        apz = str(r.get(apz_col) or "").strip()
        if not apz or apz == "-":
            continue
        saved = by_apz.get(apz)
        if not saved:
            continue
        for c in extra:
            if c in saved and (pd.isna(df.at[i, c]) or str(df.at[i, c]).strip() == ""):
                df.at[i, c] = saved[c]
    return df


def apply_hyperlinks(xlsx_path: Path, url_column_names: tuple[str, ...]) -> None:
    wb = load_workbook(xlsx_path)
    ws = wb.active
    if ws is None:
        wb.close()
        return
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col_index = {str(h): i + 1 for i, h in enumerate(header) if h}
    for name in url_column_names:
        ci = col_index.get(name)
        if not ci:
            continue
        for ri in range(2, ws.max_row + 1):
            cell = ws.cell(row=ri, column=ci)
            val = cell.value
            if isinstance(val, str) and val.startswith("http"):
                cell.hyperlink = val
                cell.font = Font(color="0563C1", underline="single")
    wb.save(xlsx_path)
    wb.close()


def write_shaffof_xlsx(df: pd.DataFrame, out_path: Path) -> Path:
    """
    Write the main sheet and URL hyperlinks. On Windows, PermissionError usually means the
    workbook is open in Excel — retry briefly, then save under a new timestamped name.
    """
    out_path = out_path.expanduser().resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    sheet = "Shaffof map"
    for delay in (0.0, 0.45, 0.95, 1.85):
        if delay:
            time.sleep(delay)
        try:
            df.to_excel(out_path, sheet_name=sheet, index=False)
            apply_hyperlinks(out_path, URL_COLS)
            return out_path
        except PermissionError:
            logging.debug("Permission denied writing %s (retry after %.2fs)", out_path, delay)
            continue
    alt = out_path.parent / f"{out_path.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{out_path.suffix}"
    logging.warning(
        "Could not write %s (permission denied). Often the file is open in Excel — close it or use "
        "the new file: %s",
        out_path,
        alt,
    )
    df.to_excel(alt, sheet_name=sheet, index=False)
    apply_hyperlinks(alt, URL_COLS)
    return alt


async def run(args: argparse.Namespace) -> int:
    setup_logging(args.verbose)
    inp = Path(args.input).expanduser().resolve()
    if not inp.exists():
        logging.error("Input not found: %s", inp)
        return 2

    df = load_input_csv(inp)
    apz_col = resolve_apz_column(df)
    for c in list(URL_COLS) + list(EXTRA_COLS):
        if c not in df.columns:
            df[c] = None

    ckpt_path = Path(args.checkpoint).expanduser().resolve() if args.checkpoint else None
    by_apz: dict[str, Any] = {}
    if ckpt_path and args.resume and ckpt_path.exists():
        raw_ck = load_checkpoint(ckpt_path)
        by_apz = dict(raw_ck.get("by_apz") or {})
        if args.retry_errors:
            before = len(by_apz)
            by_apz = {k: v for k, v in by_apz.items() if not str(v.get("scrape_error") or "").strip()}
            dropped = before - len(by_apz)
            if dropped:
                logging.info(
                    "--retry-errors: removed %s failed checkpoint entr(y/ies) so they will be scraped again",
                    dropped,
                )
                save_checkpoint(ckpt_path, by_apz)
        if args.retry_empty:
            before = len(by_apz)
            by_apz = {k: v for k, v in by_apz.items() if _ckpt_has_any_doc_url(v)}
            dropped = before - len(by_apz)
            if dropped:
                logging.info(
                    "--retry-empty: removed %s checkpoint entr(y/ies) with no document URLs so they will be scraped again",
                    dropped,
                )
                save_checkpoint(ckpt_path, by_apz)
        if args.retry_incomplete_docs:
            before = len(by_apz)
            by_apz = {k: v for k, v in by_apz.items() if _ckpt_all_four_doc_urls(v)}
            dropped = before - len(by_apz)
            if dropped:
                logging.info(
                    "--retry-incomplete-docs: removed %s checkpoint entr(y/ies) missing at least one "
                    "document URL so they will be scraped again",
                    dropped,
                )
                save_checkpoint(ckpt_path, by_apz)
        df = merge_checkpoint_into_df(df, apz_col, {"version": CKPT_VERSION, "by_apz": by_apz})
        logging.info("Resume: loaded %s completed APZ key(s) from %s", len(by_apz), ckpt_path)

    limit = int(args.limit or 0)
    processed = 0
    last_written: Path | None = None

    shot_dir = Path(args.debug_screenshots).expanduser().resolve() if args.debug_screenshots else None
    _set_run_debug(RunDebug(screenshot_dir=shot_dir, trace_actions=bool(args.trace_actions)))
    if shot_dir or args.trace_actions or args.slow_mo:
        logging.info(
            "Debug: trace_actions=%s screenshots=%s slow_mo=%sms (omit --headless for a visible window).",
            args.trace_actions,
            str(shot_dir or "off"),
            args.slow_mo,
        )

    slow_ms = max(0, int(args.slow_mo))
    use_remote = bool(args.storage) or bool(args.cdp) or bool(os.environ.get("PLAYWRIGHT_CDP_URL"))
    if use_remote and slow_ms:
        logging.warning("--slow-mo only affects bundled Chromium; ignored for --cdp / --storage.")

    try:
        async with async_playwright() as p:
            browser: Browser | None = None
            ctx: BrowserContext
            launched = False

            if args.storage:
                browser, ctx, launched = await get_browser_context(
                    p,
                    cdp=args.cdp or os.environ.get("PLAYWRIGHT_CDP_URL") or DEFAULT_CDP,
                    storage=args.storage,
                    headless=args.headless,
                    auto_chrome=not args.no_auto_chrome,
                )
            elif args.cdp or os.environ.get("PLAYWRIGHT_CDP_URL"):
                browser, ctx, launched = await get_browser_context(
                    p,
                    cdp=args.cdp or os.environ.get("PLAYWRIGHT_CDP_URL") or DEFAULT_CDP,
                    storage=None,
                    headless=args.headless,
                    auto_chrome=not args.no_auto_chrome,
                )
                if launched:
                    logging.info(
                        "Chrome was auto-started for CDP. If the map needs login, use --warmup-seconds or --storage."
                    )
            else:
                browser = await p.chromium.launch(headless=args.headless, slow_mo=slow_ms)
                ctx = await browser.new_context(locale="uz-UZ")
                logging.info("Using bundled Playwright Chromium (pass --cdp or --storage to attach elsewhere).")

            page = await ctx.new_page()
            warm = max(0, int(args.warmup_seconds))
            if warm:
                logging.info("Warmup %s s before first fetch", warm)
                await asyncio.sleep(warm)

            try:
                for i, r in df.iterrows():
                    apz = str(r.get(apz_col) or "").strip()
                    if not apz or apz == "-":
                        continue
                    if apz in by_apz and args.resume:
                        logging.info("Skip %s (checkpoint)", apz)
                        continue
                    if limit and processed >= limit:
                        logging.info("--limit %s reached", limit)
                        break

                    logging.info("APZ %s (%s/%s rows)", apz, processed + 1, len(df))
                    try:
                        result = await scrape_one_apz(page, apz)
                    except Exception as e:
                        logging.exception("Failed APZ %s", apz)
                        result = _empty_scrape_result(str(e))

                    for k, v in result.items():
                        df.at[i, k] = v
                    by_apz[apz] = {k: result.get(k) for k in list(URL_COLS) + list(EXTRA_COLS)}
                    processed += 1

                    if ckpt_path:
                        save_checkpoint(ckpt_path, by_apz)

                    out_path = Path(args.output).expanduser().resolve()
                    df_merged = merge_checkpoint_into_df(
                        df.copy(), apz_col, {"version": CKPT_VERSION, "by_apz": by_apz}
                    )
                    last_written = write_shaffof_xlsx(df_merged, out_path)

            finally:
                await page.close()
                await ctx.close()
                if browser:
                    await browser.close()

    finally:
        _set_run_debug(None)

    logging.info("Done. Wrote %s", last_written or Path(args.output).expanduser().resolve())
    return 0


def main() -> None:
    ap = argparse.ArgumentParser(description="Scrape open-data.mc.uz map URLs by APZ (from CSV).")
    ap.add_argument("--input", "-i", type=Path, default=Path("dataset.csv"))
    ap.add_argument("--output", "-o", type=Path, default=Path("shaffof_map_urls.xlsx"))
    ap.add_argument(
        "--cdp",
        type=str,
        default=None,
        help="Chrome DevTools URL (uses env PLAYWRIGHT_CDP_URL or bundled Chromium if omitted).",
    )
    ap.add_argument("--storage", type=str, default=None, help="Playwright storage state JSON path")
    ap.add_argument("--no-auto-chrome", action="store_true", help="Do not auto-start Chrome when CDP connect fails")
    ap.add_argument("--headless", action="store_true")
    ap.add_argument(
        "--trace-actions",
        action="store_true",
        help='Log "ACTION: …" for each major step (best with a visible browser, omit --headless).',
    )
    ap.add_argument(
        "--debug-screenshots",
        type=Path,
        default=None,
        metavar="DIR",
        help="Save a full-page PNG after each traced step into DIR (heavy; for debugging only).",
    )
    ap.add_argument(
        "--slow-mo",
        type=int,
        default=0,
        metavar="MS",
        help="Slow down bundled Chromium by MS ms per Playwright operation (headed debugging only).",
    )
    ap.add_argument("--checkpoint", type=Path, default=Path("shaffof_ckpt.json"))
    ap.add_argument("--resume", action="store_true", help="Skip APZ values already present in checkpoint")
    ap.add_argument(
        "--retry-errors",
        action="store_true",
        help="With --resume: remove checkpoint rows that have scrape_error so those APZ values are scraped again",
    )
    ap.add_argument(
        "--retry-empty",
        action="store_true",
        help="With --resume: remove checkpoint rows that have no http(s) document URLs (all four empty) so they are scraped again",
    )
    ap.add_argument(
        "--retry-incomplete-docs",
        action="store_true",
        help="With --resume: remove checkpoint rows missing any of the four document URL columns (not all http links) so they are scraped again",
    )
    ap.add_argument("--limit", type=int, default=0, help="Process at most N APZ rows (0 = all)")
    ap.add_argument("--warmup-seconds", type=int, default=0)
    ap.add_argument("-v", "--verbose", action="store_true")
    args = ap.parse_args()
    if args.retry_errors and not args.resume:
        ap.error("--retry-errors requires --resume")
    if args.retry_empty and not args.resume:
        ap.error("--retry-empty requires --resume")
    if args.retry_incomplete_docs and not args.resume:
        ap.error("--retry-incomplete-docs requires --resume")
    raise SystemExit(asyncio.run(run(args)))


if __name__ == "__main__":
    main()
